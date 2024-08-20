# ///////////////////////////////////////////////////////////////////////////////////////////////
# ساخت بازی حدس عدد 
# secret_number = 9 
# guess_count = 0
# guess_limit = 3
# while guess_count < guess_limit:
#     guess = int(input('Guess: ')) 
#     guess_count += 1
#     if guess == secret_number:
#         print("You Wone!")
#         break
# else:
#     print('sorry you faild')


# ///////////////////////////////////////////////////////////////////////////////////////////////
# یک مثال کاربردی برای درک حلقه فور تو در تو
# for x in range(4):
#     for y in range(3):
#         print(f'({x}, {y})')


# ///////////////////////////////////////////////////////////////////////////////////////////////
#  یک تمرین برای چاپ X

# first_way

# numbers = [5,2,5,2,2]
# for number in numbers:
#     print(number * '*')


# second_way

# numbers = [5,2,5,2,2]
# for x_count in numbers:
#     output = ''
#     for count in range(x_count):
#         output += 'X'
#     print(output)



# ///////////////////////////////////////////////////////////////////////////////////////////////
# پیدا کردن بزگترین عدد در این آرایه
# numbers = [3, 6, 2, 8, 4, 10]
# max = numbers[0]
# for number in numbers:
#     if number > max:
#         max = number
# print(max)

# ///////////////////////////////////////////////////////////////////////////////////////////////
# آرایه دو در دو Matrix 
# matrix = [
#     [1,2,3],
#     [4,5,6],
#     [7,8,9]
# ]
# print(matrix[0][1])

# for row in matrix:
#     for item in row:
#         print(item)


# ///////////////////////////////////////////////////////////////////////////////////////////////
# حذف مقدار های داپلیکیت شده در آرایه
# numbers = [2, 2, 4, 6, 3, 4, 6, 1]
# uniques = []
# for number in numbers:
#     if number not in uniques:
#         uniques.append(number)
# print(uniques)



# ///////////////////////////////////////////////////////////////////////////////////////////////
# در اینجا میخواهیم یک عددی را به عدد وارد کنیم و همان عدد به حروف نمایش داده شود
# phone = input('Phone: ')
# digits_numbers = {
#     "1" : "One",
#     "2" : "Two",
#     "3" : "Three",
#     "4" : "Four"
# }
# output = ""
# for ch in phone:
#     output += digits_numbers.get(ch , "!") + " "
# print(output)


# ///////////////////////////////////////////////////////////////////////////////////////////////
# استفاده از ایموجی در ویندوز و کد نویسی
# کافیه دکمه ویندوز + نقطه را زده تا صفحه انتخاب ایموجی باز شود
# message = input('> ')
# words = message.split(' ')
# print(words)
# emojis = {
#     ":)" : "😊",
#     ":(":"😔",
#     ":0":"😁"
# }
# output = ""
# for word in words:
#     output += emojis.get(word , word) + " "
# print(output)




# ///////////////////////////////////////////////////////////////////////////////////////////////
# یک نمرین ساده در پایتون

# sentence = input('Your Sentence: ')
# messages = sentence.split(' ')

# def print_character(message):
#     print(f'{message}')

# for message in messages:
#     print_character(message)

# ///////////////////////////////////////////////////////////////////////////////////////////////
# یک تمرین در مورد کلاس ها 
# class Person:
#     def __init__(self, name) -> None:
#         self.name = name
    
#     def talk(self):
#         print(f'talk with {self.name}')

# person = Person('amirali')
# print(person.name)
# person.talk()


# ///////////////////////////////////////////////////////////////////////////////////////////////
# استفاده از ماژول های داخل پایتون برای استفاده در کد 
# import random
# import dice

# num = dice.Dice().roll()
# print(num)

# for i in range(3):
#     print(random.random())

# for i in range(3):
#     print(random.randint(10,15))

# members = ['Amirali', 'Arian', 'Ngr', 'Seti']
# leader = random.choice(members)
# print(leader)

# ///////////////////////////////////////////////////////////////////////////////////////////////
# استفاده از ماژول های داخل پایتون برای استفاده در کد 
# استفاده از ماژول pathlib 
# from pathlib import Path

# path = Path()
# for file in path.glob('*.py'):
#     print(file)

# for file in path.glob('*'):
#     print(file)



# ///////////////////////////////////////////////////////////////////////////////////////////////
# استفاده از پکیج نصب شده در پایتون
import openpyxl as xl  

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)
# print(cell.value)
# print(sheet.max_row)
# print(sheet.max_column)

for row in range(2 , sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price
    print(cell.value)
    
wb.save('transactions2.xlsx')