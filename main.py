# //////////////////////////////////////////////////////////////////////////////////
# انواع پریمیتیو تایپ ها در پایتون و نحوه تعریف متغیر ها و نمایش آن در کنسول
# age = 20
# price = 19.5
# first_name = "Amirali"
# is_online = False
# print(age , price , first_name , is_online)


# //////////////////////////////////////////////////////////////////////////////////
#  دریافت منتغیر از ورودی و نمایش آن در کنسول
# name = input("what is your name ? ")
# print('my name is' + name)


# //////////////////////////////////////////////////////////////////////////////////
# تبدیل تایپ ها در پایتون از روش زیر استفاده میکنیم
# birth_year = input('what is your birth year ? ')
# age = 2024 - int(birth_year)
# print("my age is " + str(age))


# //////////////////////////////////////////////////////////////////////////////////
#  انواع توابع تبدیل تایپ در پایتون
# float()
# bool()
# int()
# str()


# //////////////////////////////////////////////////////////////////////////////////
# یک مثال دیگر از تبدیل تایپ در پایتون
# first = input("First :")
# second = input("Second :")
# sum: float = float(first) + float(second)
# print("Sum :" + str(sum))


# //////////////////////////////////////////////////////////////////////////////////
#  در اینجا در مورد توابع رشته ای در پایتون صحبت میکنیم
# course = "Python for beginners"
# print(course.find('y'))
# print(course.replace('for', '4'))
# print('Python' in course)  #boolean

# course = "Python for Python"
# print(course.find('y'))


# //////////////////////////////////////////////////////////////////////////////////
# arithmetic operators
# منطق ریاضیاتی در پایتون
# print(10 % 3)
# print(10 ** 3)

# x = 10
# x = x + 3
# # x += 3
# x *= 3
# print(x)

# y = 10 / 3 * 2
# print(y)



# //////////////////////////////////////////////////////////////////////////////////
# عملگر های منطقی در پایتون
# price = 15
# print(price > 10 and price < 30)
# print(price > 10 or price < 30)
# print(not price < 10)
# print(not price > 10)
#  مقادیر با True یا False نمایش داده میشوند




# //////////////////////////////////////////////////////////////////////////////////
# if statements
# شرط ها در پایتون قوانین متفاوتی دارند که در زیر میبینید
# temp = 25
# if temp < 30:
#     print('it hot day')
# elif temp > 20:
#     print('its a nice day')

# مثال
# weight : str = input("Weight: ")
# unit = input('(K) or (L)bs :')
# if unit.upper() == 'K':
#     converted = int(weight) / 0.45
#     print("Weight in Kg is" + str(converted))
# else:
#     converted = int(weight) * 0.45
#     print("Weight in Lbs is" + str(converted))




# //////////////////////////////////////////////////////////////////////////////////
#  حلقه های while
#  عدد 1000 با یک آندرلاین از هم جداشده این به ما برای خواندن بهتر عدد کمک میکند

# چاپ مستطیل
# i = 1
# while i <= 5:
#     j = 5
#     print(j * "*")
#     i += 1

# چاپ مثلث
# i = 1
# while i <= 5:
#     print(i * "*")
#     i += 1



# //////////////////////////////////////////////////////////////////////////////////

# Lists
# آرایه ها در پایتون
# names = ["John", " Bob", "Mosh", "Sam", "Marry"]
# names[0] = "jon"
# print(names)
# print(names[0])
# print(names[-2])
# print(names[0:3])


# //////////////////////////////////////////////////////////////////////////////////
#مثال ازمتد های آرایه ها
# numbers = [1, 2, 3, 4, 5,2]

# اضافه کردن عدد 20 به آخر آرایه
# numbers.append(20)


# این متد در آرایه آخرین عدد آرایه را حذف میکند
# numbers.pop(20)

# این متد مقدار یا همان متغییر را دریافت میکند و به ما ایندکس را برمیگرداند
# numbers.index(5)

# این متد آرایه ما را مرتب میکند و باید دقیقا مثل زیر استفاده شود
# numbers.sort()
# print(numbers)

# معکوس میکند
# numbers.reverse()
# print(numbers)

# یک کپی از آرایه ما میسازد که اگر آرایه اصلی تغییر کرد این تغییر نکند
# numbers2 = numbers.copy()

# این متد مقدار یا همان متغییر را دریافت میکند و به ما میگوید که چند بار تکرار شده است
# numbers.count(2)

# این متد فقط مقدار -1 را به اول آرابه اضافه میکند
# numbers.insert(0, -1)

# این متد فقط مقدار value را دریافت میکند مقدار ایندکس منظورش نیست
# numbers.remove(3)

# تمام آرایه را خالی میکند و لازم نیست هیچ ورودی به آرایه وارد کنی
# numbers.clear()

#  چک کردن اینکه این عدد در آرایه هست یا نه
# True
# print(1 in numbers)
# False
# print(7 in numbers)

#  برای بدست آوردن طول یک آرایه
# print(len(numbers))



# //////////////////////////////////////////////////////////////////////////////////
# for loop in python
# numbers = [1, 2, 3, 4]
# for item in numbers:
    # print(item)

# i = 0
# while i < len(numbers):
#     print(numbers[i])
#     i = i + 1



# //////////////////////////////////////////////////////////////////////////////////
# the range() function
# این متد یک آرایه  از 0 تا عدد 4 برای ما میسازه
#  طول این آرایه 5 خواهد بود
# numbers = range(5)
# print(numbers)
#
# for number in numbers:
#     print(number)

# این آرایه از عدد 5 شروع میشه تا عدد 9 میره
# numbers = range(5, 10)
# print(numbers)
#
# for number in numbers:
#     print(number)

# این آرایه از عدد 5 شروع میشه تا عدد 9 ولی 2 تا 2 تا پرش میکنه
# numbers = range(5, 10, 2)
# print(numbers)
#
# for number in numbers:
#     print(number)

# for number in range(5):
# print(number)

# //////////////////////////////////////////////////////////////////////////////////
# Tuples
# تاپل ها غیر قابل تغییر هستند
#  مانند لیست میمانند با این تفاوت که غیر قابلل تغییر هستند
# numbers = (1, 2, 3, 3)
# print(numbers.count(2))


# unpacking in Tuples
# coordinates = (1,2,3)
# x , y , z = coordinates
# print(x)
# print(y)
# print(z)


# //////////////////////////////////////////////////////////////////////////////////
# message
# نحوه استفاده از متغیر در استرینگ در پایتون

# name = 'john'
# last_name = 'smith'
# msg = f'{name} {last_name} is a coder'
# print(msg)


# ///////////////////////////////////////////////////////////////////////////////////////////////
# Math function
# این متد به عدد را روند میکند اگر از 2.5 به پایین بود عدد 2 و اگر از 2.5 به بالا بود عدد 3 میشه
# x = 2.6
# print(round(x))

# #قدر مطلق
# print(abs(-2.6))

# ///////////////////////////////////////////////////////////////////////////////////////////////
# ساخت بازی حدس عدد 
# secret_number = 9 
# guess_count = 0
# guess_limit = 3
# while guess_count < guess_limit:
#     guess = int(input('Guess: ')) 
#     guess_count += 1
#     if guess == secret_number:
#         print("You Wone!")
#         break
# else:
#     print('sorry you faild')


# ///////////////////////////////////////////////////////////////////////////////////////////////
# یک مثال کاربردی برای درک حلقه فور تو در تو
# for x in range(4):
#     for y in range(3):
#         print(f'({x}, {y})')


# ///////////////////////////////////////////////////////////////////////////////////////////////
#  یک تمرین برای چاپ X
# first_way
# numbers = [5,2,5,2,2]

# for number in numbers:
#     print(number * '*')

# second_way
# numbers = [5,2,5,2,2]

# for x_count in numbers:
#     output = ''
#     for count in range(x_count):
#         output += 'X'
#     print(output)


# ///////////////////////////////////////////////////////////////////////////////////////////////
# آرایه دو در دو Matrix 
# matrix = [
#     [1,2,3],
#     [4,5,6],
#     [7,8,9]
# ]
# print(matrix[0][1])

# for row in matrix:
#     for item in row:
#         print(item)

# ///////////////////////////////////////////////////////////////////////////////////////////////
# Dictionaries 
# در اینجا همان آبجکت های جاوااسکریپت را داریم با مقداری تفاوت
# customer = {
#     "name": "John Smith",
#     "age" : 30,
#     "is_verified": True
# }
# customer["name"]
# customer["age"]
# print(customer["is_verified"])
# print(customer["birthdate"])
# customer["name"] = " Jack Smith"
# print(customer)

# first way add a value to dictionaries 
# customer["birthdate"] = '10 august 1997'
# print(customer["birthdate"])

# second way add a value to dictionaries 
# print(customer.get("birthdate" , "10 august 1997"))

# ///////////////////////////////////////////////////////////////////////////////////////////////
# مقدار نان در پایتون که برابر با هیچی است
# cnumber_none = None
# print(cnumber_none)


# ///////////////////////////////////////////////////////////////////////////////////////////////
# # تعریف فانکشن در پایتون 
# def greet_user(name , last_name):
#     print(f'greet user {name} {last_name}')

# print('start')

# greet_user('Amirali','Hatami')
# greet_user(last_name='Hatami' , name='Atouse')
# greet_user('Mahnaz','Hatami')

# print('finish')


# exampleeeee

# def square(number):
#     print(number * number)
#     return None
# num =int(input('>')) 
# square(num)

# print(square(3))



# ///////////////////////////////////////////////////////////////////////////////////////////////
# Error Exception
# هندل کردن اررور ها

# try:
#     age = int(input('Age: '))
#     income = 20000
#     risk = income / age
#     print(age)
# except ZeroDivisionError:
#     print('Age Cannot be 0.')
# except ValueError:
#     print('Invalid Value')


# ///////////////////////////////////////////////////////////////////////////////////////////////
# Modules
# استفاده از ماژئل ها فقط کافیست اسم فایل را ایمپورت کنیم 
# import modules
# ایمپورت مستقیم یک تابع به یک فایل
# from modules import kg_to_lbs
# import utils 

# خواندن از یک فولدر قرار دادن فایل __init__ خیلی مهم است 
# import ecommerce.shipping
# from  ecommerce.shipping import calc_shipping

# //////
# kg_to_lbs(100)
# print(modules.kg_to_lbs(65.25))
# print(modules.lbs_to_kg(145))

# پیدا کردن بزرگترین عدد
# numbers = [10,12,3,4]
# max = utils.find_max_number(numbers)
# print(max)

# ////// 
# calc_shipping()
# ecommerce.shipping.calc_shipping()

# ///////////////////////////////////////////////////////////////////////////////////////////////
# استفاده از ماژول های داخل پایتون برای استفاده در کد 
# import random


# for i in range(3):
#     print(random.randint(10,15))
#  در فایل practice کامل هست


# ///////////////////////////////////////////////////////////////////////////////////////////////
# استفاده از path برای چک کردن اینکه فایل ها وجود دارند یا خیر
# from pathlib import Path 

# path = Path('ecommerce')
# emails = Path('emails')
# # باعث حذف دایرکتوری rmdir میشه
# emails.rmdir()
# # path_email = Path('email')
# print(path.exists())
# # print(path_email.exists())


# ///////////////////////////////////////////////////////////////////////////////////////////////
# استفاده از پکیج نصب شده در پایتون
import openpyxl as xl  

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)
print(sheet.max_column)

for row in range(1 , sheet.max_row + 1):
    cell = sheet.cell(row, 2)
    print(cell.value)

